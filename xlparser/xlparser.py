#!/usr/bin/env python3
from typing import OrderedDict,Any
import openpyxl
from openpyxl.cell import cell as cellLib
import csv
import sys
import os
from sys import argv
from datetime import datetime, timedelta, date as dateType
from dateutil.parser import parse as strptime
import json

import _io # type: ignore

__all__ = ['loadZip', 'parse', "parseXlsx", 'parseCsv',
           'parseXls', 'saveCsv', 'saveXlsx', 'openXlsx', 'rows2dict']


maxInt = sys.maxsize
while True:
    # decrease the maxInt value by factor 10 
    # as long as the OverflowError occurs.
    try:
        csv.field_size_limit(maxInt)
        break
    except OverflowError:
        maxInt = int(maxInt/10)

def loadZip(f):
    import zipfile
    import tempfile
    import os
    dest = tempfile.TemporaryDirectory()
    with open(f, 'rb') as fp:
        zf = zipfile.ZipFile(fp)
        return zf
        '''
        zf.extract('xl/media/image1.png', dest.name)
            '/var/xxxxxx/xl/media/image1.png'
        os.rename(dest.name+'/xl/media/image1.png', 'newdir')
        '''


'''
Format time
'''


def getCellValue(cell):
    debug(cell.value, type(cell.value), cell.row, cell.column,
          cell.data_type, cell.style, cell.style_id)
    # if cell.data_type in ['xxx']: # 's'
    if isinstance(cell.value, datetime):
        value = str(cell.value)
    else:
        value = cell.value
    return value


def rows2dict(rows):
    if not hasattr(rows, '__next__'):
        rows = iter(rows)
    titles = next(rows)
    for row in rows:
        yield dict(zip(titles, row))


def get_col_idx(col):
    # AZZ = 26*26+26*26+26 = 1378
    i = 0
    for char in col.encode():
        i = i*26+char-64
    return i


def openXlsx(src, active=True):
    # data_only=True, read value instead of math expression
    wb = openpyxl.load_workbook(src, read_only=False, data_only=True)
    #wb = openpyxl.load_workbook(src, read_only=False, guess_types=True)
    if active:
        return wb.active
    return wb


def getSheetSize(sh):
    try:
        max_col_idx = len(next(sh.rows))
        max_row = len(next(sh.columns))
        # if bool(sh.column_dimensions):
        #    max_col_idx = get_col_idx(list(sh.column_dimensions)[-1])
        return (max_row, max_col_idx)
    except StopIteration as e:
        return 0,0

def get_merged_cells(sh):
    l = []
    max_row, max_col_idx = getSheetSize(sh)
    for cell in sh.merged_cells.ranges:
        if cell.max_col <= max_col_idx and cell.max_row <= max_row:
            l.append(cell)
        elif cell.max_col > max_col_idx and cell.max_row > max_row:
            break
    return l


"""
isValidCell (MergedCell or Cell)
"""
def isValidCell(cell, merged_cells):
    #coord = f'{cell.column}{cell.row}'
    coord = cell.coordinate

    if cell.value != None:
        return True

    if isinstance(cell, cellLib.Cell):
        return True

    # for latest xlsx
    if isinstance(cell, openpyxl.cell.cell.MergedCell): 
        return False 

    # for old xlsx
    for i, merged_cell in enumerate(list(merged_cells)):
        if not hasattr(cell, 'col_idx'): # latest xlsx has no attribute col_idx
            return False
        # clear
        if (
                cell.col_idx > merged_cell.max_col and
                cell.row > merged_cell.max_row
        ):
            merged_cells.remove(merged_cell)

        elif (
                merged_cell.min_col <= cell.col_idx <= merged_cell.max_col and
                merged_cell.min_row <= cell.row <= merged_cell.max_row
        ):
            if (
                    merged_cell.min_col != cell.col_idx or
                    merged_cell.min_row != cell.row
            ):
                return True

    return False


def parse(src):
    if src.endswith('.xls'):
        return parseXls(src)
    if src.endswith('.xlsx'):
        return parseXlsx(src)
    if src.endswith('.csv'):
        return parseCsv(src)
    else:
        errmsg = f'Unsupported filetype:{src}'
        raise Exception(errmsg)

'''
" parse xlsx, xls
'''


def parseXlsx(src):
    wb = openpyxl.load_workbook(src, read_only=False, data_only=True)
    #sh = wb.active
    for sh in wb:
        merged_cells = get_merged_cells(sh)
        debug('merged_cells', sys._getframe().f_lineno, (merged_cells))
        rows = []
        max_row_idx, max_col_idx = getSheetSize(sh)

        debug('sh_size', max_row_idx, max_col_idx)
        for _, r in zip(range(max_row_idx), sh.rows):
            debug('--iter row-')
            row = []
            for _, cell in zip(range(max_col_idx), r):
                #if isValid(cell, merged_cells):
                v = getCellValue(cell)
                row.append(v)
            if(any(row)):
                debug(row)
                yield row

    wb.close()
    return rows


def parseXls(src):
    from xlrd import open_workbook
    wb = open_workbook(src)
    for sh in wb.sheets():
        for r in range(sh.nrows):
            row = [v if str(v).isdigit() else v for v in sh.row_values(r)]
            yield row


"""
" parseCsv
"""


def parseCsv(filep, iterator=True):
    if not isinstance(filep, _io.TextIOWrapper):
        filep = open(filep, 'r')
    rows = csv.reader(filep)
    if iterator:
        return rows
    else:
        return list(rows)


"""""""""""
saveCsv
"""""""""""


def saveCsv(rows, filep):
    # csv.writer(f, delimiter =' ',quotechar =',',quoting=csv.QUOTE_MINIMAL)
    if not isinstance(filep, _io.TextIOWrapper):
        filep = open(filep, 'w')

    try:
        c = csv.writer(filep)
        c.writerows(rows) 
    except BrokenPipeError:
        devnull = os.open(os.devnull, os.O_WRONLY)
        os.dup2(devnull, sys.stdout.fileno())
        sys.exit(1)  # Python exits with error code 1 on EPIPE


"""""""""""
formatXlsxRow
"""""""""""
def formatXlsxRow(row: list):
    return [formatXlsxCell(x) for x in row]

def formatXlsxCell(data):
    d = data if isinstance(data, (int,str,float, datetime, dateType)) else f'{data}'
    #print("format data:", type(d), d)
    return d

"""""""""""
saveXlsx
"""""""""""
def saveXlsx(rows, filep):
    if not isinstance(filep, _io.TextIOWrapper):
        filep = open(filep, 'wb')
    wb = openpyxl.Workbook()
    ws1: Any = wb.active

    if not hasattr(rows, '__next__'):
        rows = iter(rows)

    # handle first row
    try:
        row = next(rows)
    except StopIteration:
        row = None
    if isinstance(row, (dict, OrderedDict)):
        ws1.append(formatXlsxRow(row.keys()))   # type: ignore
        ws1.append(formatXlsxRow(row.values())) # type: ignore
    elif isinstance(row, (list)):
        ws1.append(formatXlsxRow(row))
    elif row==None:
        pass
    else:
        raise ValueError(f"not support type:{type(row)}, only support type dict|orderedDict|list")

    # write rows
    for row in rows:
        if isinstance(row, (dict, OrderedDict)):
            row = row.values()
        ws1.append(formatXlsxRow(row)) # type: ignore

    wb.save(filep)


if '-d' in argv:
    debug = print
else:
    debug = lambda *arg: 1
